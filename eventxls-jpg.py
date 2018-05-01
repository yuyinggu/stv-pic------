# -*- coding: utf-8 -*-
import re
import sys
from openpyxl import load_workbook, Workbook
from PIL import ImageFont, Image, ImageDraw

try:
    reload(sys)
    sys.setdefaultencoding('UTF8')
except:
    pass

# 背景图片路径
event_img_path = 'weekend_event.jpg'
# 保存图片文件夹路径
event_img_save = ''
# xlsx 文件路径
xlsx_path = 'weekend.xlsx'

################################################
# 合成活动图片
################################################

def event_to_jpg(event_arr,i):
    print(event_arr)
    try:
        im = Image.open(event_img_path)
    except Exception as e:
        print(e)
        print("活动背景图片路径不存在")
        exit()

    draw = ImageDraw.Draw(im)
    width, height = im.size
    font = ImageFont.truetype("msyhbd.ttc",48,encoding='unic')
    color = (0,0,0)
    date_width = len(re.split('\s+', event_arr[1])[0]) * 30

    draw.text((230,205),event_arr[0] ,color,font=font)
    draw.text((230,305),re.split('\s+', event_arr[1])[0],(230,3,12),font=font)
    draw.text((230+date_width,305),re.split('\s+', event_arr[1])[1] ,color,font=font)
    draw.text((230,405),event_arr[2] ,color,font=font)
    draw.text((270,565),event_arr[3] ,color,font=font)
    draw.text((270,655),event_arr[4] ,color,font=font)

    im.save(event_img_save+"%s.jpg" % str(int(i)-1), quality=100)



def main():
    wb = load_workbook('weekend.xlsx')

    sheet = wb.worksheets[0]

    row_count = sheet.max_row

    for row in range(2,row_count+1):
        row = str(row)
        event_arr = []
        event_arr.append(sheet['B'+ row].value)
        event_arr.append(sheet['C'+ row].value)
        event_arr.append(sheet['D'+ row].value)
        event_arr.append(sheet['E'+ row].value)
        event_arr.append(sheet['F'+ row].value)

        print("\n活动图片生成中...\n")
        event_to_jpg(event_arr,row)
    # print(wb[sheet_name]['A1'].value.encode('utf8'))


if __name__ == '__main__':
    main()
